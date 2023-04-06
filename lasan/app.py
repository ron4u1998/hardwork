import csv
import json
import random
import shutil
import string
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pyocr
import os
import pandas as pd
from werkzeug.utils import secure_filename
from flask import Flask, flash, redirect, render_template, request, session, send_file, send_from_directory, url_for
from PIL import Image
from appconfig import AppConfig
from pdf2image import convert_from_path
from predict import Model
import cv2
from table_predict import PredictFunc
from excel_to_json import preprocessss
from app1 import TallyPrimeIntegration

 
app = Flask(__name__)
app.config['SECRET_KEY'] = AppConfig.SECRET_KEY
############################ tesseract config ###########################################################
#pyocr.tesseract.TESSERACT_CMD = AppConfig.TESSERACT
tools = pyocr.get_available_tools()
if len(tools) == 0:
    print("No OCR tool found")
tool = tools[0]
langs = tool.get_available_languages()
lang = langs[0]

def get_text(image_path, r):
    # print(image_path)
    img = cv2.imread(image_path)
    img_crop = img[int(r[1]):int(r[1] + r[3]), int(r[0]):int(r[0] + r[2])]
    if not os.path.exists(session['TEMP']):
        os.mkdir(session['TEMP'])
    session['new_path'] = f"{session['TEMP']}/text_image.jpg"
    cv2.imwrite(session['new_path'], img_crop)
    txt = tool.image_to_string(
        Image.open(session['new_path']),
        lang=lang,
        builder=pyocr.builders.TextBuilder()
    )
    # os.remove(session['new_path'])
    print(txt)
    return txt
 

@app.route('/UploadInvoice', methods=["POST", "GET"])
def UploadInvoice():
    '''
    This function will be called
    when user clicks on dropbox to 
    upload mutiple files from local 
    directory.
    '''
    if request.files['file']:
        uploaded_file = request.files['file']
        filename = secure_filename(uploaded_file.filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext in AppConfig.UPLOAD_EXTENSIONS:
                uploaded_file.save(os.path.join(session['UPLOAD'], filename))
                
    return 'Success', 204

# SendCoordinates 
@app.route('/SendCoordinates/<filename>', methods=['GET', 'POST'])
def SendCoordinates(filename):
    '''
    This function is used to get coordinates while user 
    snips data from invoice. Coordinates will give us a small 
    portion of image from which text will be extracted.
    Coordinates with image name is sent to get_html()
    function.
    '''
    if request.method == 'POST':
        data = request.get_json()
        Coordinates = data["values"]
        if not "NewRegistrationData" in session:
            session["NewRegistrationData"] = {}
            session['Data'] = {}
            session['Counter'] = 0
            session['NewRegistrationData'][data['Snip_lableName']] = [
                Coordinates['x'], Coordinates['y'], Coordinates['w'], Coordinates['h']]
        else:
            session['NewRegistrationData'][data['Snip_lableName']] = [
                Coordinates['x'], Coordinates['y'], Coordinates['w'], Coordinates['h']]
    extractedText = get_text(os.path.join(session['OUTPUT'],filename), session['NewRegistrationData'][data['Snip_lableName']])
    session['Data'][data['Snip_lableName']] = extractedText
    return {"value": extractedText}

def UpdateJson():
    '''
    Gets CompanyName and InvoiceName from session to save it in 
    'info.json' file and excelfile with name of CompanyName 
    will be created having sheet name as InvoiceName.
    '''
    for file in session['TotalPdf']:
        wb = Workbook()
        wb.active.title = 'sheet_'+file
        wb.save(filename=os.path.join(session['EXCEL'], file)+'.xlsx')

    return 'Success', 204

def get_concat_v(im1, im2):
    dst = Image.new('RGB', (im1.width, im1.height + im2.height))
    dst.paste(im1, (0, 0))
    dst.paste(im2, (0, im1.height))
    return dst

#Extract Usefull Data From Uploaded Invoice
@app.route('/extract',methods=['GET','POST'])
def extract():
    '''
    This function will run when user clicks
    extract button. It converts pdf to image and
    passes it to model. Redirects to get_html()
    function with page and index as variable.
    '''
    if len(os.listdir(session['UPLOAD'])) == 0:
        flash(u'Please upload files','warning')
        return redirect(url_for('Registered'))
    else:
        for pdf in os.listdir(session['UPLOAD']):
            if (pdf.endswith('pdf') or pdf.endswith('PDF')):
                images = convert_from_path(os.path.join(session['UPLOAD'],pdf), dpi = 500)
                os.remove(os.path.join(session['UPLOAD'],pdf))
                for i,image in enumerate(images):
                    image.save(os.path.join(session['UPLOAD'],pdf.split('.')[0])+str(i)+'.png','PNG')
            elif (pdf.endswith('.jpg') or pdf.endswith('.JPG')):
                im1 = Image.open(os.path.join(session['UPLOAD'],pdf))
                os.remove(os.path.join(session['UPLOAD'],pdf))
                im1.save(os.path.join(session['UPLOAD'],pdf.split('.')[0]+'.png'))
                
        session['TotalPdf'] = []
        for pdf in os.listdir(session['UPLOAD']):
            session['TotalPdf'].append(pdf[:-4])
        UpdateJson()

        session['TotalPage'] = []
        for j in session['TotalPdf']:
            session['Temp'] = []
            for i in os.listdir(session['UPLOAD']):
                if(i[:-4] == j):
                    session['Temp'].append(i[:-4])
            session['TotalPage'].append(session['Temp'])

        for num in range(len(session['TotalPage'])):
            session['TotalPage'][num].sort(key=lambda x: x[-1])

        x = os.listdir(session['UPLOAD'])
        model = Model(session)
        for file in x:
            PredictFunc(file,session)
            image = cv2.imread(os.path.join(session['UPLOAD'],file))
            image = cv2.rectangle(image, (session['x'],session['y']), (session['w'],session['h']), (255, 255, 255), -1)
            session['random_string'] = ''.join(random.choice(string.ascii_letters) for _ in range(10))
            cv2.imwrite(os.path.join(session['UPLOAD'],session['random_string']+'.png'),image)
            model.predict(session['random_string']+'.png')
            model.decode(session['random_string']+'.png', file)
        return redirect(url_for('get_html',index=str(0),page=str(0)))

#Sends json data And Image Simultaneously As Per Request
@app.route('/pred/<index>/<page>',methods=['GET', 'POST'])
def get_html(index,page):
    '''
    Takes index and page which are integer values 
    and return specific file to web page.
    '''
    index = int(index)
    page  = int(page)
    book = load_workbook(os.path.join(session['SAMPLE_IMAGE'] , session['TotalPage'][index][page]+'.xlsx'))
    sheet = book.active

    session['file'] = session['TotalPage'][index][page]+'.png'
    with open(os.path.join(session['OUTPUT'],session['file'][:-4])+'.json','r') as f:
        labels = json.load(f)
    return render_template('registrationprocess.html', labels=labels, sheet=sheet , filename=session['file'], index=index, page=page, total=len(session['TotalPdf']), total_page = len(session['TotalPage'][index]))

#Display annotated image
@app.route('/display/<filename>')
def display_image(filename):
    '''
    Send requested invoice image to webpage.
    '''
    return send_from_directory(session['OUTPUT'], filename)

#send current data to tally
@app.route('/send_to_tally', methods=['GET','POST'])
def send_to_tally():
    filename = request.form.get('filename')
    index = request.form.get('index')
    page = request.form.get('page')
    # print(filename)
    preprocessss.exceltojson(os.path.join(session['SAMPLE_IMAGE'],filename.split('.')[0]+'.xlsx'))
    url = "http://localhost:9000"                
    x = TallyPrimeIntegration(url)
    x.create_ledger(os.path.join(session['OUTPUT'], filename.split('.')[0]+'.json'))
    x.fetch_stockitem(os.path.join('json_file', filename.split('.')[0]+'.json'))

    index = int(index)
    page  = int(page)
    book = load_workbook(os.path.join(session['SAMPLE_IMAGE'] , session['TotalPage'][index][page]+'.xlsx'))
    sheet = book.active

    session['file'] = session['TotalPage'][index][page]+'.png'
    with open(os.path.join(session['OUTPUT'],session['file'][:-4])+'.json','r') as f:
        labels = json.load(f)

    flash('Data sent to tally', 'success')
    return render_template('registrationprocess.html', labels=labels, sheet=sheet , filename=session['file'], index=index, page=page, total=len(session['TotalPdf']), total_page = len(session['TotalPage'][index]))



#Update fields in specific json file     
@app.route('/pred/save/<index>/<page>/<filename>', methods=['GET', 'POST'])
def saveData(index, filename, page):
    '''
    Saves data for specific filename. Takes page
    and index values to find out json files.
    '''
    di = {}
    di['INVOICE_NO']      = request.form.get('INVOICE_NO',None)
    di['INVOICE_DATE']    = request.form.get('INVOICE_DATE',None)
    di['DELIVERY_NOTE']   = request.form.get('DELIVERY_NOTE',None)
    di['DELIVERY_DATE']   = request.form.get('DELIVERY_DATE',None)
    di['COMP_NAME']       = request.form.get('COMP_NAME',None)
    di['COMP_ADD']        = request.form.get('COMP_ADD',None)
    di['COMP_CIN']        = request.form.get('COMP_CIN',None)
    di['COMP_GST']        = request.form.get('COMP_GST',None)
    di['COMP_STATE']      = request.form.get('COMP_STATE',None)
    di['COMP_STATE_CODE'] = request.form.get('COMP_STATE_CODE/BIC',None)
    di['TOTAL']           = request.form.get('TOTAL',None)
    di['TAXABLE_VALUE']   = request.form.get('TAXABLE_VALUE',None)
    di['TAX_AMOUNT']      = request.form.get('TAX_AMOUNT',None)
    with open(os.path.join(session['OUTPUT'],filename[:-4]+'.json'),'w') as f:
        json.dump(di,f)
    return redirect(url_for('get_html', index=str(index), page=str(page)))


# @app.route("/downloadAllfiles/")
# def download_all_files():
#     '''
#     This function will download zipfile. Inside zipfile
#     different excelfile with name of InvoiceType will be saved.
#     '''
#     col = mydb[session['CompanyName']]
#     x = col.find_one().keys()
#     for key in x:
#         if(key != '_id'):
#             # os.chdir(AppConfig.EXCEL_FILES)
#             wb = Workbook()
#             wb.active.title = 'sheet1'
#             wb.save(os.path.join(AppConfig.EXCEL_FILES,key+'.xlsx'))
#             x = col.find_one()[key]
#             for i in range(len(x)):
#                 data = x[i]
#                 UpdateExcelData(key,'sheet1',data)
#     file_name = session['CompanyName']+'.zip'
#     os.chdir(AppConfig.EXCEL_FILES)
#     with zipfile.ZipFile(file_name,'w') as zip:
#         for i in os.listdir():
#             if(i.endswith('.xlsx')):
#                 zip.write(i)
#     os.chdir('..')
#     return send_file(os.path.join(AppConfig.EXCEL_FILES,file_name),as_attachment=True)

def removeNone(dictionary):
    temp_dict = {}
    for pair in dictionary.items():
        if(pair[1]!='None'):
            temp_dict[pair[0]] = pair[1]
    return temp_dict


def UpdateExcelData(Data, file_name ):
    '''
    This function creates ExcelFile with CompanyName and 
    sheet with name of InvoiceType.
    '''
    df = pd.DataFrame([Data])
    OldData = pd.read_excel("{}.xlsx".format(
        session['path']), sheet_name='sheet_'+file_name)
    
    NewData = pd.concat([df, OldData])
    NewData.reset_index(drop=True)
    xl_path = session['path'] + ".xlsx"
    with pd.ExcelWriter(xl_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        NewData.to_excel(writer, sheet_name='sheet_'+file_name, index=False)
    
#Download function
@app.route("/downloadfile/", methods = ['GET','POST'])
def download_file():
    '''
    On clicking download button all JSON 
    data will be saved in single excel file
    as well as json data will be stored in 
    mongodb database inside respective company name.
    '''
    temp_dict = {
        "INVOICE_NO":       "None",
        "INVOICE_DATE":     "None", 
        "DELIVERY_NOTE":    "None", 
        "DELIVERY_DATE":    "None",
        "COMP_NAME":        "None", 
        "COMP_ADD":         "None", 
        "COMP_CIN":         "None", 
        "COMP_GST":         "None", 
        "COMP_STATE":       "None", 
        "COMP_STATE_CODE" : "None", 
        "TOTAL" :           "None", 
        "TAXABLE_VALUE" :   "None", 
        "TAX_AMOUNT" :      "None", 
        }

    for i in range(len(session['TotalPdf'])):
        for j in range(len(session['TotalPage'][i])):
            with open(os.path.join(session['OUTPUT'],session['TotalPage'][i][j]+'.json')) as json_file:
                data = json.load(json_file)
            temp_dict.update(removeNone(data))
        session['path'] = os.path.join(session['EXCEL'],session['TotalPdf'][i])
        with open(os.path.join(session['ALL_EXCEL'],session['TotalPage'][i][j]+'.json'),'w') as json_fil:
            json.dump(temp_dict, json_fil)
        UpdateExcelData(temp_dict, session['TotalPdf'][i])

    for file in os.listdir(session['SAMPLE_IMAGE']):
        if file.endswith('.xlsx'):
            df = pd.read_excel(os.path.join(session['SAMPLE_IMAGE'], file))
            df.to_json(os.path.join(session['ALL_EXCEL'],file.split('.')[0]+'_table'+'.json'), orient='index')

    session['Total_Excel'] = os.listdir(session['EXCEL'])
    session['Total_Sample_Excel'] = os.listdir(session['SAMPLE_IMAGE'])

    for i in range(len(session['TotalPdf'])):
        src_wb = load_workbook(os.path.join(session['EXCEL'],session['Total_Excel'][i]))
        # print(session['TotalPdf'])
        # print(os.path.join(session['SAMPLE_IMAGE'],session['Total_Sample_Excel'][i]))
        dest_wb = load_workbook(os.path.join(session['SAMPLE_IMAGE'],session['Total_Sample_Excel'][i]))
        session['sheet_name'] = session['Total_Excel'][i].split('.')[0]
        src_sheet = src_wb.get_sheet_by_name('sheet_'+session['sheet_name'])
        dest_sheet = dest_wb.get_sheet_by_name('Sheet1')

        for i in range(1, dest_sheet.max_row+1):
            for j in range(1, dest_sheet.max_column+1):
                if(i==1):
                    src_sheet.cell(row=i+6, column=j).font = Font(bold=True)    
                src_sheet.cell(row=i+6, column=j).value = dest_sheet.cell(row=i, column=j).value

        src_wb.save(os.path.join(session['ALL_EXCEL'],session['sheet_name']+'.xlsx'))
    for file in os.listdir(session['EXCEL']):
        if file.endswith('.xlsx'):
            df = pd.read_excel(os.path.join(session['ALL_EXCEL'],file))
            df.to_csv(os.path.join(session['ALL_EXCEL'],file.split('.')[0]+'.csv'), index=False)
            os.remove(os.path.join(session['ALL_EXCEL'],file))

    session['zip_name'] = str(session['id'])
    shutil.make_archive(os.path.join(AppConfig.ALL_EXCEL,session['zip_name']), 'zip',session['ALL_EXCEL'])

    return send_file(os.path.join(AppConfig.ALL_EXCEL,str(session['id'])+'.zip'), as_attachment=True)


#Removes all uploaded files
@app.route('/', methods=['GET'])
def Registered():
    '''
    On startup, this function will be called and 
    all directories will be emptied.
    '''
    if not session.get("id", 0):
       session['id'] = random.randint(0,50000)

    if(str(session['id']) in os.listdir(AppConfig.OUTPUT_PATH)):
        shutil.rmtree(os.path.join(AppConfig.OUTPUT_PATH,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.UPLOAD_FOLDER)):
        shutil.rmtree(os.path.join(AppConfig.UPLOAD_FOLDER,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.EXCEL_FILES)):
        shutil.rmtree(os.path.join(AppConfig.EXCEL_FILES,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.FINAL_PATH)):
        shutil.rmtree(os.path.join(AppConfig.FINAL_PATH,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.TEMP_DIR)):
        shutil.rmtree(os.path.join(AppConfig.TEMP_DIR,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.SAMPLE_IMAGE)):
        shutil.rmtree(os.path.join(AppConfig.SAMPLE_IMAGE,str(session['id'])))
    if(str(session['id']) in os.listdir(AppConfig.ALL_EXCEL)):
        shutil.rmtree(os.path.join(AppConfig.ALL_EXCEL,str(session['id'])))

    session['UPLOAD'] = os.path.join(AppConfig.UPLOAD_FOLDER, str(session['id']))
    session['OUTPUT'] = os.path.join(AppConfig.OUTPUT_PATH, str(session['id']))
    session['FINAL'] = os.path.join(AppConfig.FINAL_PATH, str(session['id']))
    session['EXCEL'] = os.path.join(AppConfig.EXCEL_FILES, str(session['id']))
    session['TEMP'] = os.path.join(AppConfig.TEMP_DIR, str(session['id']))
    session['SAMPLE_IMAGE'] = os.path.join(AppConfig.SAMPLE_IMAGE, str(session['id']))
    session['ALL_EXCEL'] = os.path.join(AppConfig.ALL_EXCEL, str(session['id']))

    os.mkdir(session['UPLOAD'])
    os.mkdir(session['OUTPUT'])
    os.mkdir(session['FINAL'])
    os.mkdir(session['EXCEL'])
    os.mkdir(session['TEMP'])
    os.mkdir(session['SAMPLE_IMAGE'])
    os.mkdir(session['ALL_EXCEL'])

    if('info.json' not in os.listdir()):
        with open('info.json','w') as f :
            data = {"CompanyName":''}
            json.dump(data,f)
        OutputArray = [[]]
        return render_template('registeredinvoiceextraction.html',Company_Name=OutputArray)
    else:
        f = open('info.json')
        data = json.load(f)
        f.close()
        OutputArray = []
        for key,values in data.items():
            outputObj = {'brand_id': values,'brand_name': values}
            OutputArray.append(outputObj)

    return render_template('registeredinvoiceextraction.html', Company_Name=OutputArray)

@app.route("/Registered/InvoiceType", methods=["POST", "GET"])
def get_company_name():
    '''
    Get CompanyName from html file and return
    list of InvoiceName of respective company.
    '''
    if request.method == 'POST':
        category_id = request.form['category_id']
        session['CompanyName'] =  category_id
        f = open('info.json')
        data = json.load(f)
        f.close()
        if (session['CompanyName'] not in data.values()):
            data['CompanyName_'+str(len(data.values()))] = session['CompanyName']
        with open('info.json','w') as json_file:
            json.dump(data,json_file)
        
    return 'Success', 204

if __name__=='__main__':
    app.run(host='0.0.0.0', port=5006)
