from openpyxl import load_workbook
from openpyxl.styles import Font

path1 = r'123.xlsx'
path2 = r'1234.xlsx'
def merge_excel(path1, path2):
    src_wb = load_workbook(path1)
    dest_wb = load_workbook(path2)

    src_sheet = src_wb.get_sheet_by_name('sheet_123')
    dest_sheet = dest_wb.get_sheet_by_name('Sheet1')

    for i in range(1, dest_sheet.max_row+1):
        for j in range(1, dest_sheet.max_column+1):
            print(dest_sheet.cell(row=i, column=j).value)
            Font(bold=True)
            if(i==1):
                src_sheet.cell(row=i+6, column=j).font = Font(bold=True)    
            src_sheet.cell(row=i+6, column=j).value = dest_sheet.cell(row=i, column=j).value

    src_wb.save('source.xlsx')

x = merge_excel(path1, path2)














